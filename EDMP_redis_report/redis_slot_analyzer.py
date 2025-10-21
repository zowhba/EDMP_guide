#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Redis 슬롯 분석 프로그램
기존 Excel 파일을 열어서 stb_id를 기반으로 Redis 슬롯을 계산하고 
G, H 컬럼을 추가한 후 새로운 파일로 저장하는 프로그램

사용법:
python redis_slot_analyzer.py <excel_file_path>
"""

import sys
import os
import pandas as pd
import hashlib
from datetime import datetime
import argparse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def calculate_redis_slot(stb_id):
    """
    stb_id를 기반으로 Redis 슬롯 번호를 계산합니다.
    중괄호를 제거한 stb_id를 키로 사용하여 CRC16 해시를 계산합니다.
    Redis Cluster는 CRC16 해시를 사용하여 16384개의 슬롯 중 하나를 선택합니다.
    """
    if pd.isna(stb_id) or stb_id == "":
        return ""
    
    # stb_id를 문자열로 변환하고 중괄호 제거
    key = str(stb_id).strip('{}').strip()
    
    # CRC16 계산 (Redis Cluster에서 사용하는 방식)
    def crc16(data):
        """CRC16 해시 계산"""
        crc = 0
        for byte in data.encode('utf-8'):
            crc ^= byte << 8
            for _ in range(8):
                if crc & 0x8000:
                    crc = (crc << 1) ^ 0x1021
                else:
                    crc <<= 1
                crc &= 0xFFFF
        return crc
    
    # Redis 슬롯 계산 (0-16383)
    slot = crc16(key) % 16384
    return slot


def get_redis_server(slot_number):
    """
    슬롯 번호를 기반으로 Redis 서버 정보를 반환합니다.
    주어진 Excel 수식을 Python으로 구현
    """
    if slot_number == "" or pd.isna(slot_number):
        return "error"
    
    try:
        slot = int(slot_number)
        if slot <= 2730:
            return "redis 1"
        elif slot <= 5460:
            return "redis 2"
        elif slot <= 8191:
            return "redis 3"
        elif slot <= 10922:
            return "redis 4"
        elif slot <= 13652:
            return "redis 5"
        elif slot <= 16383:
            return "redis 6"
        else:
            return "error"
    except (ValueError, TypeError):
        return "error"


def apply_header_styling(excel_file_path):
    """
    Excel 파일의 헤더 행에 노란색 음영을 적용하고 컬럼 너비를 자동 조정하며, Redis Node 컬럼에 수식을 추가합니다.
    """
    try:
        # Excel 파일 열기
        workbook = load_workbook(excel_file_path)
        worksheet = workbook.active
        
        # 노란색 배경 스타일 정의
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # 첫 번째 행(헤더)의 모든 셀에 노란색 배경 적용
        for cell in worksheet[1]:  # 첫 번째 행
            if cell.value:  # 값이 있는 셀만
                cell.fill = yellow_fill
        
        # Redis Slot과 Redis Node 컬럼 찾기
        redis_slot_col = None
        redis_node_col = None
        
        for col_idx, cell in enumerate(worksheet[1], 1):
            if cell.value == 'Redis Slot':
                redis_slot_col = col_idx
            elif cell.value == 'Redis Node':
                redis_node_col = col_idx
        
        # Redis Node 컬럼에 Excel 수식 추가
        if redis_slot_col and redis_node_col:
            print("📝 Redis Node 컬럼에 Excel 수식을 추가하는 중...")
            
            # 데이터 행 수 계산 (헤더 제외)
            max_row = worksheet.max_row
            
            for row_idx in range(2, max_row + 1):  # 2행부터 시작 (헤더 제외)
                # Redis Slot 컬럼의 셀 참조 (예: G2, G3, ...)
                slot_cell_ref = f"{worksheet.cell(row=1, column=redis_slot_col).column_letter}{row_idx}"
                
                # Excel 수식 생성
                formula = f'=IF({slot_cell_ref}="", "error", IF({slot_cell_ref}<=2730, "redis 1", IF({slot_cell_ref}<=5460, "redis 2", IF({slot_cell_ref}<=8191, "redis 3", IF({slot_cell_ref}<=10922, "redis 4", IF({slot_cell_ref}<=13652, "redis 5", IF({slot_cell_ref}<=16383, "redis 6", "error")))))))'
                
                # Redis Node 셀에 수식 설정
                worksheet.cell(row=row_idx, column=redis_node_col).value = formula
        
        # 모든 컬럼의 너비를 내용에 맞춰 자동 조정
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        # 셀 값의 길이 계산 (한글은 2배로 계산)
                        cell_length = len(str(cell.value))
                        # 한글 문자가 포함된 경우 추가 공간 확보
                        if any('\uac00' <= char <= '\ud7af' for char in str(cell.value)):
                            cell_length = int(cell_length * 1.2)  # 한글 고려하여 20% 추가
                        
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # 최소 너비 8, 최대 너비 50으로 제한
            adjusted_width = min(max(max_length + 2, 8), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 파일 저장
        workbook.save(excel_file_path)
        workbook.close()
        
    except Exception as e:
        print(f"⚠️  스타일링 중 오류 발생: {str(e)}")
        print("   (데이터는 정상적으로 저장되었습니다)")


def process_excel_file(input_file_path):
    """
    기존 Excel 파일을 열어서 Redis 슬롯 정보를 추가한 후 새 파일로 저장합니다.
    """
    try:
        # Excel 파일 읽기
        print(f"📂 기존 Excel 파일을 읽는 중: {input_file_path}")
        df = pd.read_excel(input_file_path)
        
        # 기존 파일 정보 출력
        print(f"📊 기존 파일 정보:")
        print(f"   - 행 수: {len(df)}")
        print(f"   - 컬럼 수: {len(df.columns)}")
        print(f"   - 기존 컬럼: {list(df.columns)}")
        
        # stb_id 컬럼이 있는지 확인
        if 'stb_id' not in df.columns:
            print("❌ 오류: 'stb_id' 컬럼을 찾을 수 없습니다.")
            print(f"사용 가능한 컬럼: {list(df.columns)}")
            return None
        
        # 기존 Redis 관련 컬럼 확인
        if 'Redis Slot' in df.columns:
            print("⚠️  기존 'Redis Slot' 컬럼이 존재합니다. 덮어쓰기됩니다.")
        if 'Redis Node' in df.columns:
            print("⚠️  기존 'Redis Node' 컬럼이 존재합니다. 덮어쓰기됩니다.")
        
        print(f"\n🔄 총 {len(df)} 행의 데이터를 처리합니다.")
        
        # 기존 컬럼과 새 컬럼 사이에 빈 열 추가
        print("📋 기존 컬럼과 새 컬럼 사이에 구분용 빈 열을 추가하는 중...")
        df[''] = ''  # 빈 컬럼 추가
        
        # Redis Slot 컬럼: Redis 슬롯 계산
        print("🔢 'Redis Slot' 컬럼에 Redis 슬롯을 계산하는 중...")
        df['Redis Slot'] = df['stb_id'].apply(calculate_redis_slot)
        
        # Redis Node 컬럼: Excel 수식으로 설정 (나중에 openpyxl로 처리)
        print("🖥️  'Redis Node' 컬럼을 준비하는 중...")
        df['Redis Node'] = ''  # 빈 값으로 초기화 (나중에 수식으로 대체)
        
        # 결과 파일명 생성
        input_dir = os.path.dirname(input_file_path)
        today = datetime.now().strftime("%Y-%m-%d")
        output_filename = f"Redis Timeout분석결과(발생일자 {today}).xlsx"
        output_path = os.path.join(input_dir, output_filename)
        
        # 결과 저장 (먼저 pandas로 저장)
        print(f"💾 새로운 파일로 저장하는 중: {output_path}")
        df.to_excel(output_path, index=False)
        
        # 헤더 스타일링, 컬럼 너비 자동 조정, Excel 수식 추가
        print("🎨 헤더 스타일링, 컬럼 너비 조정, Excel 수식 추가 중...")
        apply_header_styling(output_path)
        
        # 처리 결과 요약
        slot_counts = df['Redis Slot'].value_counts().sort_index()
        
        # Redis Node는 Excel 수식으로 처리되므로 Python에서 미리 계산해서 분포 표시
        redis_node_preview = df['Redis Slot'].apply(get_redis_server).value_counts()
        
        print("\n✅ === 처리 완료 ===")
        print(f"📁 입력 파일: {input_file_path}")
        print(f"📁 출력 파일: {output_path}")
        print(f"📊 처리된 행 수: {len(df)}")
        print(f"📊 최종 컬럼 수: {len(df.columns)} (Redis Slot, Redis Node 컬럼 추가됨)")
        print(f"📝 Redis Node 컬럼: Excel 수식으로 설정됨")
        print("\n📈 === Redis 서버별 분포 (예상) ===")
        for server, count in redis_node_preview.items():
            print(f"   {server}: {count}개")
        
        return output_path
        
    except FileNotFoundError:
        print(f"오류: 파일을 찾을 수 없습니다 - {input_file_path}")
        return None
    except Exception as e:
        print(f"오류 발생: {str(e)}")
        return None


def main():
    """메인 함수"""
    parser = argparse.ArgumentParser(
        description="Redis 슬롯 분석 프로그램",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
사용 예시:
  python redis_slot_analyzer.py data.xlsx
  python redis_slot_analyzer.py /path/to/data.xlsx
        """
    )
    
    parser.add_argument(
        'excel_file',
        help='분석할 Excel 파일 경로'
    )
    
    args = parser.parse_args()
    
    # 파일 경로 확인
    if not os.path.exists(args.excel_file):
        print(f"오류: 파일이 존재하지 않습니다 - {args.excel_file}")
        sys.exit(1)
    
    # 파일 처리
    result = process_excel_file(args.excel_file)
    
    if result:
        print(f"\n✅ 성공적으로 완료되었습니다!")
        print(f"결과 파일: {result}")
        sys.exit(0)
    else:
        print("\n❌ 처리 중 오류가 발생했습니다.")
        sys.exit(1)


if __name__ == "__main__":
    main()
